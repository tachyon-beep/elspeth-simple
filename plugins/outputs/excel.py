"""Result sink that materialises experiment outputs into an Excel workbook."""

from __future__ import annotations

import json
import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, Mapping

from dmp.core.interfaces import ResultSink, Artifact, ArtifactDescriptor
from dmp.core.security import normalize_security_level


logger = logging.getLogger(__name__)


def _load_workbook_dependencies():  # type: ignore[return-any]
    try:
        from openpyxl import Workbook  # type: ignore
    except ImportError as exc:  # pragma: no cover - handled during sink initialisation
        raise RuntimeError(
            "ExcelResultSink requires the 'openpyxl' package. Install with 'pip install openpyxl'"
        ) from exc
    return Workbook


class ExcelResultSink(ResultSink):
    """Persist experiment payloads into a timestamped Excel workbook."""

    def __init__(
        self,
        *,
        base_path: str | Path,
        workbook_name: str | None = None,
        timestamped: bool = True,
        results_sheet: str = "Results",
        manifest_sheet: str = "Manifest",
        aggregates_sheet: str = "Aggregates",
        include_manifest: bool = True,
        include_aggregates: bool = True,
        on_error: str = "abort",
    ) -> None:
        self.base_path = Path(base_path)
        self.workbook_name = workbook_name
        self.timestamped = timestamped
        self.results_sheet = results_sheet
        self.manifest_sheet = manifest_sheet
        self.aggregates_sheet = aggregates_sheet
        self.include_manifest = include_manifest
        self.include_aggregates = include_aggregates
        if on_error not in {"abort", "skip"}:
            raise ValueError("on_error must be 'abort' or 'skip'")
        self.on_error = on_error
        # Ensure dependency availability early for fast failure when configured incorrectly.
        self._workbook_factory = _load_workbook_dependencies()
        self._last_workbook_path: str | None = None
        self._security_level: str | None = None

    # ------------------------------------------------------------------ public API
    def write(self, results: Dict[str, Any], *, metadata: Dict[str, Any] | None = None) -> None:
        metadata = metadata or {}
        timestamp = datetime.now(timezone.utc)
        try:
            path = self._resolve_path(metadata, timestamp)
            path.parent.mkdir(parents=True, exist_ok=True)

            workbook = self._workbook_factory()
            self._populate_results_sheet(workbook, results.get("results", []))

            if self.include_manifest:
                self._populate_manifest_sheet(workbook, results, metadata, timestamp)

            if self.include_aggregates and results.get("aggregates"):
                self._populate_aggregates_sheet(workbook, results.get("aggregates"))

            workbook.save(path)
            self._last_workbook_path = str(path)
            if metadata:
                self._security_level = normalize_security_level(metadata.get("security_level"))
        except Exception as exc:
            if self.on_error == "skip":
                logger.warning("Excel sink failed; skipping workbook creation: %s", exc)
                return
            raise

    # ------------------------------------------------------------------ helpers
    def _resolve_path(self, metadata: Mapping[str, Any], timestamp: datetime) -> Path:
        name = self.workbook_name or str(
            metadata.get("experiment") or metadata.get("name") or "experiment"
        )
        if name.endswith(".xlsx"):
            name = name.removesuffix(".xlsx")
        if self.timestamped:
            name = f"{name}_{timestamp.strftime('%Y%m%dT%H%M%SZ')}"
        return self.base_path / f"{name}.xlsx"

    def _populate_results_sheet(self, workbook, entries: Iterable[Mapping[str, Any]]) -> None:  # type: ignore[no-untyped-def]
        sheet = workbook.active
        sheet.title = self.results_sheet

        flattened = [self._flatten_result(entry) for entry in entries]
        headers: list[str] = []
        if flattened:
            headers = sorted({key for row in flattened for key in row.keys()})
            sheet.append(headers)
            for row in flattened:
                sheet.append([row.get(column) for column in headers])
        else:
            sheet.append(["no_results"])

    def _populate_manifest_sheet(
        self,
        workbook,
        results: Mapping[str, Any],
        metadata: Mapping[str, Any],
        timestamp: datetime,
    ) -> None:  # type: ignore[no-untyped-def]
        sheet = workbook.create_sheet(self.manifest_sheet)
        manifest = self._build_manifest(results, metadata, timestamp)
        sheet.append(["key", "value"])
        for key, value in manifest.items():
            if isinstance(value, (dict, list)):
                sheet.append([key, json.dumps(value, sort_keys=True)])
            else:
                sheet.append([key, value])

    def _populate_aggregates_sheet(self, workbook, aggregates: Mapping[str, Any]) -> None:  # type: ignore[no-untyped-def]
        sheet = workbook.create_sheet(self.aggregates_sheet)
        sheet.append(["metric", "value"])
        for key, value in aggregates.items():
            if isinstance(value, Mapping):
                sheet.append([key, json.dumps(value, sort_keys=True)])
            else:
                sheet.append([key, value])

    @staticmethod
    def _flatten_result(entry: Mapping[str, Any]) -> Dict[str, Any]:
        flat: Dict[str, Any] = {}
        row = entry.get("row")
        if isinstance(row, Mapping):
            for key, value in row.items():
                flat[f"row.{key}"] = value
        for key, value in entry.items():
            if key == "row":
                continue
            if isinstance(value, (dict, list)):
                flat[key] = json.dumps(value, sort_keys=True)
            else:
                flat[key] = value
        return flat

    @staticmethod
    def _build_manifest(
        results: Mapping[str, Any],
        metadata: Mapping[str, Any],
        timestamp: datetime,
    ) -> Dict[str, Any]:
        manifest = {
            "generated_at": timestamp.isoformat(),
            "rows": len(results.get("results", [])) if isinstance(results.get("results"), list) else 0,
            "metadata": dict(metadata),
        }
        if "cost_summary" in results:
            manifest["cost_summary"] = results["cost_summary"]
        if "failures" in results:
            manifest["failures"] = results["failures"]
        return manifest

    def produces(self):  # pragma: no cover - placeholder for artifact chaining
        return [
            ArtifactDescriptor(name="excel", type="file/xlsx", persist=True, alias="excel"),
        ]

    def consumes(self):  # pragma: no cover - placeholder for artifact chaining
        return []

    def finalize(self, artifacts, *, metadata=None):  # pragma: no cover - optional cleanup
        return None

    def collect_artifacts(self) -> Dict[str, Artifact]:  # pragma: no cover
        if not self._last_workbook_path:
            return {}
        artifact = Artifact(
            id="",
            type="file/xlsx",
            path=self._last_workbook_path,
            metadata={
                "path": self._last_workbook_path,
                "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "security_level": self._security_level,
            },
            persist=True,
            security_level=self._security_level,
        )
        self._last_workbook_path = None
        self._security_level = None
        return {"excel": artifact}


__all__ = ["ExcelResultSink"]
